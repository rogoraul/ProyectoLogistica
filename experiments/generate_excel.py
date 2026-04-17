"""
generate_excel.py
Genera 3 ficheros Excel con los resultados de la experimentación:
  1. calibracion_grasp.xlsx  — calibración del parámetro alpha para GRASP
  2. calibracion_ts.xlsx     — calibración de alpha y tenure para GRASP+TS
  3. comparacion_final.xlsx  — comparación final GRASP vs GRASP+TS

Ejecutar desde la raíz del proyecto:
    python experiments/generate_excel.py
"""

import csv
import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# ─────────────────────────────────────────────────────────────────────────────
# Paleta y estilos
# ─────────────────────────────────────────────────────────────────────────────
C_HDR_BG    = "1F3864"  # azul marino
C_HDR_FG    = "FFFFFF"
C_SUB_BG    = "2E75B6"  # azul medio
C_SUB_FG    = "FFFFFF"
C_ALT       = "DCE6F1"  # azul muy claro (filas pares)
C_WHITE     = "FFFFFF"
C_BEST      = "E2EFDA"  # verde claro → mejor config
C_WORST     = "FCE4D6"  # naranja claro → peor / challenger
C_BORDER    = "BDD7EE"
FONT        = "Arial"


def _thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(color):
    return PatternFill("solid", start_color=color)


def _hdr(cell, text, bg=C_HDR_BG, fg=C_HDR_FG, size=9):
    cell.value = text
    cell.font = Font(name=FONT, bold=True, color=fg, size=size)
    cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _dat(cell, value, even_row, fmt=None, align="center", bold=False):
    cell.value = value
    cell.font = Font(name=FONT, size=9, bold=bold)
    cell.fill = _fill(C_ALT if even_row else C_WHITE)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _thin_border()
    if fmt:
        cell.number_format = fmt


def _highlight(cell, color):
    cell.fill = _fill(color)


def _col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title(ws, title, subtitle, start_row=1):
    """Escribe bloque de título de 2 filas y devuelve la primera fila de datos."""
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=12)
    c = ws.cell(start_row, 1, title)
    c.font = Font(name=FONT, bold=True, size=13, color=C_HDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 24

    ws.merge_cells(start_row=start_row+1, start_column=1,
                   end_row=start_row+1, end_column=12)
    c2 = ws.cell(start_row+1, 1, subtitle)
    c2.font = Font(name=FONT, size=8, italic=True, color="595959")
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row+1].height = 14
    return start_row + 3   # primera fila libre


def _sub_banner(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    _hdr(ws.cell(row, 1), text, bg=C_SUB_BG)
    ws.row_dimensions[row].height = 18


def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def read_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def alpha_label(v):
    try:
        return "Aleatorio" if float(v) == -1 else v
    except Exception:
        return v


# ─────────────────────────────────────────────────────────────────────────────
# 1. calibracion_grasp.xlsx
# ─────────────────────────────────────────────────────────────────────────────

def build_grasp_excel():
    summary = read_csv(os.path.join(HERE, "calibration_grasp_summary.csv"))
    detail  = read_csv(os.path.join(HERE, "calibration_grasp.csv"))
    runs    = read_csv(os.path.join(HERE, "calibration_grasp_runs.csv"))

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen por Alpha ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen por Alpha"
    fdr = _title(ws1,
                 "CALIBRACIÓN GRASP – Resumen por valor de Alpha",
                 "Alpha controla el balance entre voracidad y aleatoriedad  "
                 "| Tiempo límite: 10 s  | 3 ejecuciones por configuración")

    hdrs = ["Grupo", "Alpha", "Dev. Media (%)", "OF Media", "Tiempo Medio (s)", "Ranking"]
    for c, h in enumerate(hdrs, start=1):
        _hdr(ws1.cell(fdr, c), h)
    ws1.row_dimensions[fdr].height = 28
    ws1.freeze_panes = f"A{fdr+1}"

    # Determinar mejor dev% por grupo
    best_dev = {}
    for r in summary:
        g = r["group"]
        d = float(r["avg_dev_pct"])
        if g not in best_dev or d < best_dev[g]:
            best_dev[g] = d

    # Ordenar: grupo + dev%
    from functools import cmp_to_key
    def sort_key(r):
        return (r["group"], float(r["avg_dev_pct"]))
    summary_sorted = sorted(summary, key=sort_key)

    # Añadir ranking por grupo
    rank_counter = {}
    for r in summary_sorted:
        g = r["group"]
        rank_counter[g] = rank_counter.get(g, 0) + 1
        r["_rank"] = rank_counter[g]

    row = fdr + 1
    prev_g = None
    for r in summary_sorted:
        if r["group"] != prev_g:
            label = ("Instancias Pequeñas (n=100, p=10)"
                     if r["group"] == "small"
                     else "Instancias Grandes (n=500, p=50)")
            _sub_banner(ws1, row, label, len(hdrs))
            row += 1
            prev_g = r["group"]

        even = (row % 2 == 0)
        vals = [r["group"], alpha_label(r["alpha"]),
                float(r["avg_dev_pct"])/100,
                float(r["mean_avg_of"]), float(r["mean_time_s"]), r["_rank"]]
        fmts = [None, None, "0.0000%", "#,##0.00", "0.000", "0"]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws1.cell(row, c), v, even, fmt)
        if float(r["avg_dev_pct"]) == best_dev[r["group"]]:
            for c in range(1, len(hdrs)+1):
                _highlight(ws1.cell(row, c), C_BEST)
        row += 1

    row += 1
    ws1.cell(row, 1, "Verde = mejor configuración por grupo  |  Dev.% = desviación respecto al mejor valor encontrado").font = \
        Font(name=FONT, size=8, italic=True, color="595959")

    _col_widths(ws1, [20, 12, 16, 14, 18, 10])

    # ── Hoja 2: Detalle por Instancia ────────────────────────────────────
    ws2 = wb.create_sheet("Detalle por Instancia")
    fdr2 = _title(ws2,
                  "CALIBRACIÓN GRASP – Detalle por instancia y alpha",
                  "Estadísticas agregadas de 3 runs por cada combinación instancia × alpha")

    hdrs2 = ["Grupo", "Instancia", "Alpha", "OF Media", "OF Mejor",
             "OF Peor", "Desv. Est.", "Tiempo (s)", "Dev. Media (%)"]
    for c, h in enumerate(hdrs2, start=1):
        _hdr(ws2.cell(fdr2, c), h)
    ws2.row_dimensions[fdr2].height = 28
    ws2.freeze_panes = f"A{fdr2+1}"

    detail_sorted = sorted(detail, key=lambda r: (r["group"], r["instance"], float(r["alpha"])))
    for i, r in enumerate(detail_sorted, start=1):
        row2 = fdr2 + i
        even = (i % 2 == 0)
        vals = [r["group"], r["instance"], alpha_label(r["alpha"]),
                float(r["avg_of"]), float(r["best_of"]), float(r["worst_of"]),
                float(r["std_of"]), float(r["avg_time_s"]), float(r["avg_dev_pct"])/100]
        fmts = [None, None, None, "#,##0.00", "#,##0.00", "#,##0.00",
                "#,##0.00", "0.000", "0.0000%"]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws2.cell(row2, c), v, even, fmt,
                 align="left" if c == 2 else "center")

    _col_widths(ws2, [14, 30, 10, 12, 12, 12, 12, 12, 14])

    # ── Hoja 3: Runs Individuales ─────────────────────────────────────────
    ws3 = wb.create_sheet("Runs Individuales")
    fdr3 = _title(ws3,
                  "CALIBRACIÓN GRASP – Ejecuciones individuales",
                  "Cada fila es una ejecución independiente  |  Semilla base = 42  (semilla_i = 42 + i)")

    hdrs3 = ["Grupo", "Instancia", "Alpha", "Run", "Semilla", "OF", "Tiempo (s)"]
    for c, h in enumerate(hdrs3, start=1):
        _hdr(ws3.cell(fdr3, c), h)
    ws3.row_dimensions[fdr3].height = 28
    ws3.freeze_panes = f"A{fdr3+1}"
    ws3.auto_filter.ref = f"A{fdr3}:{get_column_letter(len(hdrs3))}{fdr3}"

    for i, r in enumerate(runs, start=1):
        row3 = fdr3 + i
        even = (i % 2 == 0)
        vals = [r["group"], r["instance"], alpha_label(r["alpha"]),
                int(r["run"]), int(r["seed"]), float(r["of"]), float(r["elapsed_s"])]
        fmts = [None, None, None, "0", "0", "#,##0.00", "0.000"]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws3.cell(row3, c), v, even, fmt,
                 align="left" if c == 2 else "center")

    _col_widths(ws3, [14, 30, 10, 8, 10, 12, 12])

    out = os.path.join(HERE, "calibracion_grasp.xlsx")
    wb.save(out)
    print(f"[OK] {out}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. calibracion_ts.xlsx
# ─────────────────────────────────────────────────────────────────────────────

def build_ts_excel():
    summary = read_csv(os.path.join(HERE, "calibration_ts_summary.csv"))
    detail  = read_csv(os.path.join(HERE, "calibration_ts.csv"))
    runs    = read_csv(os.path.join(HERE, "calibration_ts_runs.csv"))

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen por Configuración ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen por Configuración"
    fdr = _title(ws1,
                 "CALIBRACIÓN GRASP+TS – Resumen de configuraciones",
                 "Proceso secuencial: Fase 1 = barrido alpha (tenure=15 fijo)  "
                 "| Fase 2 = barrido tenure (mejor alpha fijo)  | 10 s / 3 runs por config.")

    hdrs = ["Grupo", "Fase", "Alpha", "Tenure", "Dev. Media (%)", "OF Media", "Tiempo (s)", "Ranking"]
    for c, h in enumerate(hdrs, start=1):
        _hdr(ws1.cell(fdr, c), h)
    ws1.row_dimensions[fdr].height = 28
    ws1.freeze_panes = f"A{fdr+1}"

    best_dev = {}
    for r in summary:
        key = (r["group"], r["phase"])
        d = float(r["avg_dev_pct"])
        if key not in best_dev or d < best_dev[key]:
            best_dev[key] = d

    summary_sorted = sorted(summary, key=lambda r: (r["group"], r["phase"], float(r["avg_dev_pct"])))
    rank_counter = {}
    for r in summary_sorted:
        key = (r["group"], r["phase"])
        rank_counter[key] = rank_counter.get(key, 0) + 1
        r["_rank"] = rank_counter[key]

    row = fdr + 1
    prev_key = None
    for r in summary_sorted:
        key = (r["group"], r["phase"])
        if key != prev_key:
            grp  = "Pequeñas (n=100)" if r["group"] == "small" else "Grandes (n=500)"
            fase = ("Fase 1 – Barrido Alpha (tenure=15)"
                    if r["phase"] == "alpha_sweep"
                    else "Fase 2 – Barrido Tenure (alpha fijado)")
            _sub_banner(ws1, row, f"{grp}  ·  {fase}", len(hdrs))
            row += 1
            prev_key = key

        even = (row % 2 == 0)
        fase_str = "Barrido Alpha" if r["phase"] == "alpha_sweep" else "Barrido Tenure"
        vals = [r["group"], fase_str, alpha_label(r["alpha"]), int(r["tenure"]),
                float(r["avg_dev_pct"])/100, float(r["mean_avg_of"]),
                float(r["mean_time_s"]), r["_rank"]]
        fmts = [None, None, None, "0", "0.0000%", "#,##0.00", "0.000", "0"]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws1.cell(row, c), v, even, fmt)
        if float(r["avg_dev_pct"]) == best_dev[key]:
            for c in range(1, len(hdrs)+1):
                _highlight(ws1.cell(row, c), C_BEST)
        row += 1

    row += 1
    ws1.cell(row, 1, "Verde = mejor configuración por grupo y fase").font = \
        Font(name=FONT, size=8, italic=True, color="595959")
    _col_widths(ws1, [14, 18, 10, 10, 16, 14, 12, 10])

    # ── Hoja 2: Fase 1 – Barrido Alpha ───────────────────────────────────
    ws2 = wb.create_sheet("Fase 1 – Barrido Alpha")
    fdr2 = _title(ws2,
                  "CALIBRACIÓN GRASP+TS – Fase 1: Barrido de Alpha",
                  "Tenure fijo = 15  |  Valores de alpha probados: 0.1, 0.25, 0.5, 0.75, 0.9, Aleatorio")

    hdrs2 = ["Grupo", "Instancia", "Alpha", "Tenure", "OF Media", "OF Mejor",
             "OF Peor", "Desv. Est.", "Tiempo (s)", "Dev. Media (%)"]
    for c, h in enumerate(hdrs2, start=1):
        _hdr(ws2.cell(fdr2, c), h)
    ws2.row_dimensions[fdr2].height = 28
    ws2.freeze_panes = f"A{fdr2+1}"

    alpha_sweep = [r for r in detail if r["phase"] == "alpha_sweep"]
    alpha_sweep_sorted = sorted(alpha_sweep,
                                key=lambda r: (r["group"], r["instance"], float(r["alpha"])))
    for i, r in enumerate(alpha_sweep_sorted, start=1):
        row2 = fdr2 + i
        even = (i % 2 == 0)
        vals = [r["group"], r["instance"], alpha_label(r["alpha"]), int(r["tenure"]),
                float(r["avg_of"]), float(r["best_of"]), float(r["worst_of"]),
                float(r["std_of"]), float(r["avg_time_s"]), float(r["avg_dev_pct"])/100]
        fmts = [None, None, None, "0",
                "#,##0.00", "#,##0.00", "#,##0.00", "#,##0.00", "0.000", "0.0000%"]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws2.cell(row2, c), v, even, fmt,
                 align="left" if c == 2 else "center")

    _col_widths(ws2, [14, 30, 10, 10, 12, 12, 12, 12, 12, 14])

    # ── Hoja 3: Fase 2 – Barrido Tenure ──────────────────────────────────
    ws3 = wb.create_sheet("Fase 2 – Barrido Tenure")
    fdr3 = _title(ws3,
                  "CALIBRACIÓN GRASP+TS – Fase 2: Barrido de Tenure",
                  "Alpha fijado al mejor de Fase 1  |  Tenures probados: 5, 10, 15, 20, 25, 30")

    hdrs3 = ["Grupo", "Instancia", "Alpha", "Tenure", "OF Media", "OF Mejor",
             "OF Peor", "Desv. Est.", "Tiempo (s)", "Dev. Media (%)"]
    for c, h in enumerate(hdrs3, start=1):
        _hdr(ws3.cell(fdr3, c), h)
    ws3.row_dimensions[fdr3].height = 28
    ws3.freeze_panes = f"A{fdr3+1}"

    tenure_sweep = [r for r in detail if r["phase"] == "tenure_sweep"]
    tenure_sweep_sorted = sorted(tenure_sweep,
                                 key=lambda r: (r["group"], r["instance"], int(r["tenure"])))
    for i, r in enumerate(tenure_sweep_sorted, start=1):
        row3 = fdr3 + i
        even = (i % 2 == 0)
        vals = [r["group"], r["instance"], alpha_label(r["alpha"]), int(r["tenure"]),
                float(r["avg_of"]), float(r["best_of"]), float(r["worst_of"]),
                float(r["std_of"]), float(r["avg_time_s"]), float(r["avg_dev_pct"])/100]
        fmts = [None, None, None, "0",
                "#,##0.00", "#,##0.00", "#,##0.00", "#,##0.00", "0.000", "0.0000%"]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws3.cell(row3, c), v, even, fmt,
                 align="left" if c == 2 else "center")

    _col_widths(ws3, [14, 30, 10, 10, 12, 12, 12, 12, 12, 14])

    # ── Hoja 4: Runs Individuales ─────────────────────────────────────────
    ws4 = wb.create_sheet("Runs Individuales")
    fdr4 = _title(ws4,
                  "CALIBRACIÓN GRASP+TS – Ejecuciones individuales",
                  "Cada fila es una ejecución independiente  |  Semilla base = 42  (semilla_i = 42 + i)")

    hdrs4 = ["Grupo", "Fase", "Instancia", "Alpha", "Tenure",
             "Run", "Semilla", "OF", "Tiempo (s)"]
    for c, h in enumerate(hdrs4, start=1):
        _hdr(ws4.cell(fdr4, c), h)
    ws4.row_dimensions[fdr4].height = 28
    ws4.freeze_panes = f"A{fdr4+1}"
    ws4.auto_filter.ref = f"A{fdr4}:{get_column_letter(len(hdrs4))}{fdr4}"

    for i, r in enumerate(runs, start=1):
        row4 = fdr4 + i
        even = (i % 2 == 0)
        fase_label = ("Barrido Alpha" if r["phase"] == "alpha_sweep" else "Barrido Tenure")
        vals = [r["group"], fase_label, r["instance"], alpha_label(r["alpha"]),
                int(r["tenure"]), int(r["run"]), int(r["seed"]),
                float(r["of"]), float(r["elapsed_s"])]
        fmts = [None, None, None, None, "0", "0", "0", "#,##0.00", "0.000"]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws4.cell(row4, c), v, even, fmt,
                 align="left" if c == 3 else "center")

    _col_widths(ws4, [14, 16, 30, 10, 10, 8, 10, 12, 12])

    out = os.path.join(HERE, "calibracion_ts.xlsx")
    wb.save(out)
    print(f"[OK] {out}")


# ─────────────────────────────────────────────────────────────────────────────
# 3. comparacion_final.xlsx
# ─────────────────────────────────────────────────────────────────────────────

def build_comparison_excel():
    results = read_csv(os.path.join(HERE, "comparison_results.csv"))
    runs    = read_csv(os.path.join(HERE, "comparison_runs.csv"))
    tests   = read_json(os.path.join(HERE, "comparison_tests.json"))

    small_res = [r for r in results if r["group"] == "small"]
    large_res = [r for r in results if r["group"] == "large"]

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen Ejecutivo ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    fdr = _title(ws1,
                 "COMPARACIÓN FINAL – GRASP vs GRASP+TS",
                 "15 instancias (6 pequeñas + 9 grandes)  |  30 s por ejecución  |  5 runs por instancia")

    def kv_row(ws, row, label, value, note=""):
        c_l = ws.cell(row, 1, label)
        c_l.font = Font(name=FONT, bold=True, size=9, color=C_HDR_BG)
        c_l.alignment = Alignment(horizontal="left", vertical="center")
        c_l.border = _thin_border()
        c_v = ws.cell(row, 2, value)
        c_v.font = Font(name=FONT, size=9)
        c_v.alignment = Alignment(horizontal="center", vertical="center")
        c_v.border = _thin_border()
        if note:
            c_n = ws.cell(row, 3, note)
            c_n.font = Font(name=FONT, size=8, italic=True, color="595959")

    row = fdr
    _sub_banner(ws1, row, "Parámetros de experimentación", 4)
    row += 1
    params = [
        ("Tiempo límite por run", "30 s"),
        ("Runs por instancia", "5"),
        ("Semilla base", "42  (42 + i por run)"),
        ("Alpha GRASP – pequeñas / grandes", "0.1 / 0.1"),
        ("Alpha GRASP+TS – pequeñas / grandes", "0.25 / 0.9"),
        ("Tenure GRASP+TS – pequeñas / grandes", "15 / 10"),
        ("Test estadístico", "Wilcoxon signed-rank (bilateral, p < 0.05)"),
    ]
    for label, val in params:
        kv_row(ws1, row, label, val)
        row += 1

    row += 1
    _sub_banner(ws1, row, "Métricas globales", 4)
    row += 1
    kv_row(ws1, row, "N.º instancias totales", len(results)); row += 1
    kv_row(ws1, row, "N.º instancias pequeñas (n=100)", len(small_res)); row += 1
    kv_row(ws1, row, "N.º instancias grandes (n=500)", len(large_res)); row += 1

    row += 1
    _sub_banner(ws1, row, "Tabla comparativa – Instancias Grandes (n=500, p=50)", 4)
    row += 1

    comp_hdrs = ["Métrica", "GRASP", "GRASP+TS", "Ganador"]
    for c, h in enumerate(comp_hdrs, start=1):
        _hdr(ws1.cell(row, c), h)
    ws1.row_dimensions[row].height = 22
    row += 1

    g_dev  = sum(float(r["grasp_dev_pct"]) for r in large_res) / len(large_res)
    ts_dev = sum(float(r["ts_dev_pct"])    for r in large_res) / len(large_res)
    g_of   = sum(float(r["grasp_avg_of"]) for r in large_res) / len(large_res)
    ts_of  = sum(float(r["ts_avg_of"])    for r in large_res) / len(large_res)
    g_std  = sum(float(r["grasp_std"])    for r in large_res) / len(large_res)
    ts_std = sum(float(r["ts_std"])       for r in large_res) / len(large_res)
    g_nbk  = sum(int(r["grasp_nbest"])    for r in large_res)
    ts_nbk = sum(int(r["ts_nbest"])       for r in large_res)

    comp_rows = [
        ("Dev. media (%) sobre mejor conocido",  g_dev/100,  ts_dev/100,
         "GRASP" if g_dev < ts_dev else "GRASP+TS",  "0.0000%"),
        ("OF media",                              g_of,       ts_of,
         "GRASP" if g_of > ts_of else "GRASP+TS",    "#,##0.00"),
        ("Desviación estándar media",             g_std,      ts_std,
         "GRASP" if g_std < ts_std else "GRASP+TS",  "#,##0.00"),
        ("Veces que alcanza el mejor conocido",   g_nbk,      ts_nbk,
         "GRASP" if g_nbk > ts_nbk else "GRASP+TS",  "0"),
        ("p-valor Wilcoxon (instancias grandes)", 0.002628,  "—",
         "Diferencia significativa (α=0.05)",          "0.000000"),
    ]

    for i, (label, gv, tv, winner, fmt) in enumerate(comp_rows, start=1):
        even = (i % 2 == 0)
        c_l = ws1.cell(row, 1, label)
        c_l.font = Font(name=FONT, size=9)
        c_l.fill = _fill(C_ALT if even else C_WHITE)
        c_l.alignment = Alignment(horizontal="left", vertical="center")
        c_l.border = _thin_border()

        for col, v in [(2, gv), (3, tv)]:
            cell = ws1.cell(row, col)
            cell.value = v
            cell.font = Font(name=FONT, size=9)
            cell.fill = _fill(C_ALT if even else C_WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()
            if isinstance(v, (int, float)):
                cell.number_format = fmt

        c_w = ws1.cell(row, 4, winner)
        c_w.font = Font(name=FONT, size=9, bold=True)
        c_w.fill = _fill(C_BEST)
        c_w.alignment = Alignment(horizontal="center", vertical="center")
        c_w.border = _thin_border()
        row += 1

    row += 1
    _sub_banner(ws1, row, "Conclusión", 4)
    row += 1
    conclusion = (
        "GRASP es más consistente (menor Dev.% media y menor varianza). "
        "GRASP+TS tiene mayor varianza pero puede encontrar mejores soluciones puntuales. "
        "La diferencia es estadísticamente significativa (Wilcoxon p=0.0026)."
    )
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c_conc = ws1.cell(row, 1, conclusion)
    c_conc.font = Font(name=FONT, size=9, italic=True)
    c_conc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws1.row_dimensions[row].height = 40

    _col_widths(ws1, [42, 16, 16, 24])

    # ── Hoja 2: Resultados por Instancia ─────────────────────────────────
    ws2 = wb.create_sheet("Resultados por Instancia")
    fdr2 = _title(ws2,
                  "COMPARACIÓN FINAL – Resultados por instancia",
                  "Estadísticas de 5 runs por instancia  |  Verde = OF media mayor  |  Naranja = OF media menor")

    hdrs2 = [
        "Instancia", "Grupo", "Mejor Conocido",
        "GRASP α", "GRASP OF Media", "GRASP Mejor", "GRASP Peor", "GRASP Desv.", "GRASP Dev.%", "GRASP #BK", "GRASP T(s)",
        "TS α", "TS Tenure", "TS OF Media", "TS Mejor", "TS Peor", "TS Desv.", "TS Dev.%", "TS #BK", "TS T(s)",
    ]
    for c, h in enumerate(hdrs2, start=1):
        _hdr(ws2.cell(fdr2, c), h, size=8)
    ws2.row_dimensions[fdr2].height = 36
    ws2.freeze_panes = f"A{fdr2+1}"

    results_sorted = sorted(results, key=lambda r: (r["group"], r["instance"]))
    for i, r in enumerate(results_sorted, start=1):
        row2 = fdr2 + i
        even = (i % 2 == 0)
        vals = [
            r["instance"], r["group"], float(r["best_known"]),
            float(r["grasp_alpha"]),
            float(r["grasp_avg_of"]), float(r["grasp_best"]), float(r["grasp_worst"]),
            float(r["grasp_std"]), float(r["grasp_dev_pct"])/100,
            int(r["grasp_nbest"]), float(r["grasp_avg_time"]),
            float(r["ts_alpha"]), int(r["ts_tenure"]),
            float(r["ts_avg_of"]), float(r["ts_best"]), float(r["ts_worst"]),
            float(r["ts_std"]), float(r["ts_dev_pct"])/100,
            int(r["ts_nbest"]), float(r["ts_avg_time"]),
        ]
        fmts = [
            None, None, "#,##0.00",
            "0.00", "#,##0.00", "#,##0.00", "#,##0.00", "#,##0.00", "0.0000%", "0", "0.000",
            "0.00", "0", "#,##0.00", "#,##0.00", "#,##0.00", "#,##0.00", "0.0000%", "0", "0.000",
        ]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws2.cell(row2, c), v, even, fmt,
                 align="left" if c == 1 else "center")

        g_of_val  = float(r["grasp_avg_of"])
        ts_of_val = float(r["ts_avg_of"])
        if g_of_val > ts_of_val:
            _highlight(ws2.cell(row2, 5), C_BEST)
            _highlight(ws2.cell(row2, 14), C_WORST)
        elif ts_of_val > g_of_val:
            _highlight(ws2.cell(row2, 5), C_WORST)
            _highlight(ws2.cell(row2, 14), C_BEST)

    _col_widths(ws2, [28, 10, 14, 8, 12, 12, 12, 12, 10, 8, 10,
                      8, 8, 12, 12, 12, 12, 10, 8, 10])

    # ── Hoja 3: Instancias Pequeñas ───────────────────────────────────────
    ws3 = wb.create_sheet("Instancias Pequeñas")
    fdr3 = _title(ws3,
                  "COMPARACIÓN – Instancias Pequeñas (n=100, p=10)",
                  "6 instancias  |  Ambos algoritmos alcanzan el óptimo en todas las ejecuciones  |  Dev.% = 0.00%")

    hdrs3 = ["Instancia", "Mejor Conocido",
             "GRASP OF Media", "GRASP Dev.%", "GRASP #BK",
             "TS OF Media", "TS Dev.%", "TS #BK", "Resultado"]
    for c, h in enumerate(hdrs3, start=1):
        _hdr(ws3.cell(fdr3, c), h)
    ws3.row_dimensions[fdr3].height = 28
    ws3.freeze_panes = f"A{fdr3+1}"

    for i, r in enumerate(sorted(small_res, key=lambda x: x["instance"]), start=1):
        row3 = fdr3 + i
        even = (i % 2 == 0)
        g_of_val  = float(r["grasp_avg_of"])
        ts_of_val = float(r["ts_avg_of"])
        resultado = ("Empate" if g_of_val == ts_of_val else
                     "GRASP" if g_of_val > ts_of_val else "GRASP+TS")
        vals = [r["instance"], float(r["best_known"]),
                g_of_val, float(r["grasp_dev_pct"])/100, int(r["grasp_nbest"]),
                ts_of_val, float(r["ts_dev_pct"])/100, int(r["ts_nbest"]),
                resultado]
        fmts = [None, "#,##0.00", "#,##0.00", "0.0000%", "0",
                "#,##0.00", "0.0000%", "0", None]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws3.cell(row3, c), v, even, fmt,
                 align="left" if c == 1 else "center")
        if resultado == "Empate":
            _highlight(ws3.cell(row3, 9), C_ALT)

    _col_widths(ws3, [28, 14, 14, 12, 8, 14, 12, 8, 12])

    # ── Hoja 4: Instancias Grandes ────────────────────────────────────────
    ws4 = wb.create_sheet("Instancias Grandes")
    fdr4 = _title(ws4,
                  "COMPARACIÓN – Instancias Grandes (n=500, p=50)",
                  "9 instancias  |  GRASP más consistente  |  GRASP+TS mayor varianza  |  p-Wilcoxon=0.0026")

    hdrs4 = ["Instancia", "Mejor Conocido",
             "GRASP OF Media", "GRASP Dev.%", "GRASP Desv.", "GRASP #BK",
             "TS OF Media", "TS Dev.%", "TS Desv.", "TS #BK",
             "Ganador OF Media"]
    for c, h in enumerate(hdrs4, start=1):
        _hdr(ws4.cell(fdr4, c), h)
    ws4.row_dimensions[fdr4].height = 28
    ws4.freeze_panes = f"A{fdr4+1}"

    for i, r in enumerate(sorted(large_res, key=lambda x: x["instance"]), start=1):
        row4 = fdr4 + i
        even = (i % 2 == 0)
        g_of_val  = float(r["grasp_avg_of"])
        ts_of_val = float(r["ts_avg_of"])
        ganador = ("Empate" if g_of_val == ts_of_val else
                   "GRASP" if g_of_val > ts_of_val else "GRASP+TS")
        vals = [r["instance"], float(r["best_known"]),
                g_of_val, float(r["grasp_dev_pct"])/100, float(r["grasp_std"]), int(r["grasp_nbest"]),
                ts_of_val, float(r["ts_dev_pct"])/100, float(r["ts_std"]), int(r["ts_nbest"]),
                ganador]
        fmts = [None, "#,##0.00",
                "#,##0.00", "0.0000%", "#,##0.00", "0",
                "#,##0.00", "0.0000%", "#,##0.00", "0", None]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws4.cell(row4, c), v, even, fmt,
                 align="left" if c == 1 else "center")
        c_w = ws4.cell(row4, 11)
        if ganador == "GRASP":
            _highlight(c_w, C_BEST)
        elif ganador == "GRASP+TS":
            _highlight(c_w, C_WORST)
        else:
            _highlight(c_w, C_ALT)

    foot_row = fdr4 + len(large_res) + 2
    ws4.cell(foot_row, 1,
             "Verde = GRASP gana  |  Naranja = GRASP+TS gana  "
             "|  Wilcoxon global p=0.002628  →  significativo al 95%").font = \
        Font(name=FONT, size=8, italic=True, color="595959")

    _col_widths(ws4, [28, 14, 14, 10, 12, 8, 14, 10, 12, 8, 14])

    # ── Hoja 5: Test Estadístico ──────────────────────────────────────────
    ws5 = wb.create_sheet("Test Estadístico")
    fdr5 = _title(ws5,
                  "TEST ESTADÍSTICO – Wilcoxon Signed-Rank",
                  "Pareado por semilla (misma semilla → misma instancia → mismo algoritmo)  |  Bilateral  |  α=0.05")

    hdrs5 = ["Grupo", "N Pares", "Wins GRASP", "Wins GRASP+TS", "Empates",
             "Delta medio (TS−GRASP)", "Pares no-cero", "Estadístico W", "p-valor", "Conclusión"]
    for c, h in enumerate(hdrs5, start=1):
        _hdr(ws5.cell(fdr5, c), h)
    ws5.row_dimensions[fdr5].height = 28
    ws5.freeze_panes = f"A{fdr5+1}"

    group_order = [("overall", "Global"), ("small", "Pequeñas"), ("large", "Grandes")]
    for i, (key, label) in enumerate(group_order, start=1):
        t = tests[key]
        row5 = fdr5 + i
        even = (i % 2 == 0)
        if t["status"] == "ok":
            conclusion = ("Significativo (p<0.05)" if t["pvalue"] < 0.05
                          else "No significativo (p≥0.05)")
            vals = [label, t["n_pairs"], t["wins_grasp"], t["wins_ts"], t["ties"],
                    t["mean_delta_ts_minus_grasp"],
                    t.get("non_zero_pairs", "-"), t.get("statistic", "-"),
                    t["pvalue"], conclusion]
            fmts = [None, "0", "0", "0", "0", "#,##0.00", "0", "#,##0.00", "0.000000", None]
        else:
            reason = t.get("reason", t["status"])
            vals = [label, t["n_pairs"],
                    t.get("wins_grasp", "-"), t.get("wins_ts", "-"), t.get("ties", "-"),
                    t.get("mean_delta_ts_minus_grasp", "-"),
                    "-", "-", "-", f"Omitido: {reason}"]
            fmts = [None, "0", None, None, None, None, None, None, None, None]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws5.cell(row5, c), v, even, fmt)
        if t.get("status") == "ok" and t.get("pvalue", 1) < 0.05:
            _highlight(ws5.cell(row5, 9), C_BEST)
            _highlight(ws5.cell(row5, 10), C_BEST)

    note_row = fdr5 + len(group_order) + 2
    ws5.cell(note_row, 1,
             "Nota: test de dos colas, zero_method='wilcox'  "
             "|  Verde = p < 0.05 (diferencia estadísticamente significativa)").font = \
        Font(name=FONT, size=8, italic=True, color="595959")

    _col_widths(ws5, [14, 10, 12, 14, 10, 20, 12, 16, 12, 28])

    # ── Hoja 6: Runs Individuales ─────────────────────────────────────────
    ws6 = wb.create_sheet("Runs Individuales")
    fdr6 = _title(ws6,
                  "COMPARACIÓN FINAL – Ejecuciones individuales",
                  "150 filas (15 instancias × 2 algoritmos × 5 runs)  |  Base para el test de Wilcoxon")

    hdrs6 = ["Instancia", "Grupo", "Algoritmo", "Run", "Semilla",
             "Alpha", "Tenure", "OF", "Tiempo (s)"]
    for c, h in enumerate(hdrs6, start=1):
        _hdr(ws6.cell(fdr6, c), h)
    ws6.row_dimensions[fdr6].height = 28
    ws6.freeze_panes = f"A{fdr6+1}"
    ws6.auto_filter.ref = f"A{fdr6}:{get_column_letter(len(hdrs6))}{fdr6}"

    runs_sorted = sorted(runs, key=lambda r: (r["group"], r["instance"], r["algorithm"], int(r["run"])))
    for i, r in enumerate(runs_sorted, start=1):
        row6 = fdr6 + i
        even = (i % 2 == 0)
        tenure_val = "" if not r["tabu_tenure"] else int(r["tabu_tenure"])
        vals = [r["instance"], r["group"], r["algorithm"],
                int(r["run"]), int(r["seed"]), float(r["alpha"]),
                tenure_val, float(r["of"]), float(r["elapsed_s"])]
        fmts = [None, None, None, "0", "0", "0.00",
                "0" if tenure_val != "" else None,
                "#,##0.00", "0.000"]
        for c, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            _dat(ws6.cell(row6, c), v, even, fmt,
                 align="left" if c in (1, 3) else "center")

    _col_widths(ws6, [28, 10, 12, 8, 10, 10, 10, 12, 12])

    out = os.path.join(HERE, "comparacion_final.xlsx")
    wb.save(out)
    print(f"[OK] {out}")


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    build_grasp_excel()
    build_ts_excel()
    build_comparison_excel()
    print("\nTodos los ficheros Excel generados correctamente.")
